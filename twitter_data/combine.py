""" 本程序是filter.py的后继程序，将筛选出来的推文里，转发自同一推文的进行合并 """
""" 相同推文的判断标准是，推文A的10%处到60%处，这一段完整存在于推文B中，且A与Bc长度差异不超过20%，则A、B认为是相同推文 """
import os
import sys
import jsonlines
import json
from openpyxl import Workbook

current_path = os.getcwd()

path = "xinjiang_cotton"

filedir = os.path.join(current_path, path)
filenames = os.listdir(filedir)
target_file = os.path.join(current_path, path + ".xlsx")

all_tweets = {}
index = 0
for filename in filenames:
    print (filename)
    filepath = os.path.join(filedir, filename)
    with open (filepath, "r") as f:
        for tweets in jsonlines.Reader(f):
            tweets["filename"] = filename
            # tweets["flag"] = False
            all_tweets[index] = tweets
            index += 1

print(len(all_tweets))
# sys.exit(0)

record_the_use_of_tweet = all_tweets.copy()
output = []
# used_tweet = []
for i in range(len(all_tweets)):
    # if all_tweets[i]["flag"] == False:
    # if all_tweets[i]["id_str"] not in used_tweet:
    # if i not in used_tweet:
    if i in record_the_use_of_tweet:
        print(all_tweets[i]["filename"], all_tweets[i]["id_str"])
        count = 0
        if all_tweets[i]["text"].startswith("RT") == False:
            try:
                try:
                    text_original = all_tweets[i]["extended_tweet"]["full_text"]
                except:
                    text_original = all_tweets[i]["text"]
            except:
                print("ERROR! No Tweet:", all_tweets[i]["filename"], all_tweets[i]["id_str"])
                sys.exit(0)
            output.append([all_tweets[i]["filename"], all_tweets[i]["created_at"], all_tweets[i]["user"]["screen_name"], all_tweets[i]["id_str"], text_original, count])
            # all_tweets[i]["flag"] = True
            # used_tweet.append(all_tweets[i]["id_str"])
            # used_tweet.append(i)
            record_the_use_of_tweet.pop(i)
            # print("Pop: i =", i)
            len_original = len(text_original)
            for j in range(i+1, len(all_tweets)):
                if j in record_the_use_of_tweet:
                    try:
                        try:
                            text_j = all_tweets[j]["extended_tweet"]["full_text"]
                        except:
                            try:
                                text_j = all_tweets[j]["retweeted_status"]["extended_tweet"]["full_text"]
                            except:
                                text_j = all_tweets[j]["text"]
                    except:
                        # 如果在对应位置找不到推文，即上述try的内容出错，那意味着肯定不是同一个推文的retweet，直接看下一个推文即可
                        continue
                    if text_original[round(0.1*len_original): round(0.6*len_original)] in text_j and len(text_original) > 0.9*len(text_j) and len(text_original) < 1.1*len(text_j):
                        count += 1
                        output.append([all_tweets[j]["filename"], all_tweets[j]["created_at"], all_tweets[j]["user"]["screen_name"], all_tweets[j]["id_str"], text_j, count])
                        # all_tweets[j]["flag"] = True
                        # used_tweet.append(all_tweets[j]["id_str"])
                        # used_tweet.append(j)
                        record_the_use_of_tweet.pop(j)
                        # print("Pop: j =", j)
        else:
            try:
                try:
                    text_original = all_tweets[i]["retweeted_status"]["extended_tweet"]["full_text"]
                except:
                    text_original = all_tweets[i]["text"]
            except:
                print("ERROR! No Tweet:", all_tweets[i]["filename"], all_tweets[i]["id_str"])
                sys.exit(0)
            output.append([all_tweets[i]["filename"], all_tweets[i]["created_at"], all_tweets[i]["user"]["screen_name"], all_tweets[i]["id_str"], text_original, count])
            # all_tweets[i]["flag"] = True
            # used_tweet.append(all_tweets[i]["id_str"])
            # used_tweet.append(i)
            record_the_use_of_tweet.pop(i)
            # print("Pop: i =", i)
            len_original = len(text_original)
            for j in range(i+1, len(all_tweets)):
                if j in record_the_use_of_tweet and all_tweets[j]["text"].startswith("RT") == True:
                    try:
                        try:
                            text_j = all_tweets[j]["retweeted_status"]["extended_tweet"]["full_text"]
                        except:
                            text_j = all_tweets[j]["text"]
                    except:
                        # 如果在对应位置找不到推文，即上述try的内容出错，那意味着肯定不是同一个推文的retweet，直接看下一个推文即可
                        continue
                    if text_original[round(0.1*len_original): round(0.6*len_original)] in text_j and len(text_original) > 0.9*len(text_j) and len(text_original) < 1.1*len(text_j):
                        count += 1
                        output.append([all_tweets[j]["filename"], all_tweets[j]["created_at"], all_tweets[j]["user"]["screen_name"], all_tweets[j]["id_str"], text_j, count])
                        # all_tweets[j]["flag"] = True
                        # used_tweet.append(all_tweets[j]["id_str"])
                        # used_tweet.append(j)
                        record_the_use_of_tweet.pop(j)
                        # print("Pop: j =", j)

length_tweets = len(output)
print("Number of Tweets: ", length_tweets)
wb = Workbook()
ws = wb.active
ws["A1"] = "filename"
ws["B1"] = "created_at"
ws["C1"] = "user"
ws["D1"] = "id_str"
ws["E1"] = "text"
ws["F1"] = "number_retweeted"
for i in range(length_tweets):
    ws.append(output[i])
wb.save(target_file)
